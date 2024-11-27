import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from scipy.signal import find_peaks
from openpyxl import Workbook

# Sélection du côté, 'R' pour droit, 'L' pour gauche
side_selector = 'L'  # 'R' ou 'L'

# Dictionnaire de paramètres selon le côté
params = {
    'R': {
        'Side_name': 'Right',
        'x_Ankle': 'x_Right_Ankle',
        'y_Ankle': 'y_Right_Ankle',
        'x_Knee': 'x_Right_Knee',
        'y_Knee': 'y_Right_Knee',
        'x_Hip': 'x_Right_Hip',
        'y_Hip': 'y_Right_Hip',
        'x_Shoulder': 'x_Right_Shoulder',
        'y_Shoulder': 'y_Right_Shoulder',
        'x_Toe': 'x_Right_big_Toe',
        'y_Toe': 'y_Right_big_Toe',
        'opposite_side_displacement': 'x_Left_Ankle',
        'opposite_side_name': 'Left'
    },
    'L': {
        'Side_name': 'Left',
        'x_Ankle': 'x_Left_Ankle',
        'y_Ankle': 'y_Left_Ankle',
        'x_Knee': 'x_Left_Knee',
        'y_Knee': 'y_Left_Knee',
        'x_Hip': 'x_Left_Hip',
        'y_Hip': 'y_Left_Hip',
        'x_Shoulder': 'x_Left_Shoulder',
        'y_Shoulder': 'y_Left_Shoulder',
        'x_Toe': 'x_Left_big_Toe',
        'y_Toe': 'y_Left_big_Toe',
        'opposite_side_displacement': 'x_Right_Ankle',
        'opposite_side_name': 'Right'
    }
}


# Définition des vecteurs et calcul de l'angle entre eux
def vector_angle(v1, v2):
    # produit scalaire :
    dot_product = np.dot(v1, v2)
    # calcule le produit des normes des 2 vecteurs
    magnitude_product = np.linalg.norm(v1) * np.linalg.norm(v2)
    # Évite la division par zéro
    if magnitude_product == 0:
        return np.nan
    # Calcul du cosinus de l'angle (produit scalaire divisé par le produit des normes)
    cos_theta = dot_product / magnitude_product
    return np.arccos(np.clip(cos_theta, -1.0, 1.0))


# Calcul des angles de la hanche, du genou et de la cheville
def calculate_angles(df, side_params):
    angles = []
    for index, row in df.iterrows():
        # Vérification des colonnes nécessaires
        required_columns = [
            side_params['x_Ankle'],
            side_params['x_Knee'],
            side_params['y_Ankle'],
            side_params['y_Knee'],
            side_params['x_Hip'],
            side_params['y_Hip'],
            side_params['x_Shoulder'],
            side_params['y_Shoulder'],
            side_params['x_Toe'],
            side_params['y_Toe']
        ]

        # Vérifie si toutes les colonnes nécessaires sont présentes
        if not all(col in df.columns for col in required_columns):
            print(
                f"Colonnes manquantes pour le côté {side_params['Side_name']}: "
                f"{[col for col in required_columns if col not in df.columns]}")
            return None

        # Vecteurs pour le genou
        ankle_to_knee = np.array([
            row[side_params['x_Ankle']] - row[side_params['x_Knee']],
            row[side_params['y_Ankle']] - row[side_params['y_Knee']]
        ])
        knee_to_hip = np.array([
            row[side_params['x_Knee']] - row[side_params['x_Hip']],
            row[side_params['y_Knee']] - row[side_params['y_Hip']]
        ])

        # Vecteurs pour la hanche
        hip_to_shoulder = np.array([
            row[side_params['x_Hip']] - row[side_params['x_Shoulder']],
            row[side_params['y_Hip']] - row[side_params['y_Shoulder']]
        ])

        # Vecteurs pour la cheville
        toe_to_ankle = np.array([
            row[side_params['x_Toe']] - row[side_params['x_Ankle']],
            row[side_params['y_Toe']] - row[side_params['y_Ankle']]
        ])

        # Calcul des angles en radians
        knee_angle = vector_angle(ankle_to_knee, knee_to_hip)  # Angle du genou
        hip_angle = vector_angle(knee_to_hip, hip_to_shoulder)  # Angle de la hanche
        ankle_angle = vector_angle(ankle_to_knee, toe_to_ankle)  # Angle de la cheville

        # Ajout des angles à la liste
        angles.append([knee_angle, hip_angle, ankle_angle])

    # Conversion en degrés
    angles_deg = np.degrees(angles)
    return pd.DataFrame(angles_deg, columns=['Knee_Angle', 'Hip_Angle', 'Ankle_Angle'])


# Trace les graphes des angles et les enregistre sous forme de fichiers PNG
def plot_angle(angle_data, angle_name, output_file):
    plt.figure(figsize=(8, 6))
    plt.plot(angle_data, label=angle_name, color='blue')
    plt.title(f"{angle_name} over Time")
    plt.xlabel("Frame")
    plt.ylabel("Angle (degrees)")
    plt.grid()
    plt.savefig(output_file)
    plt.show()


# Fonction de traçage du déplacement en x de la cheville
def plot_displacement(df, x_ankle_mvt, side_params):
    plt.figure(figsize=(8, 6))
    plt.plot(df[side_params['opposite_side_displacement']], label=side_params['opposite_side_name'], color='purple')
    plt.title(f"Displacement of {side_params['Side_name']} Ankle (x) over Time")
    plt.xlabel("Frame")
    plt.ylabel("x Position")
    plt.grid()
    plt.savefig(x_ankle_mvt)
    plt.show()


# Méthode pour trouver les pics et vallées, puis calculer la différence
def calculate_horizontal_displacement(df, side_params):
    # Lissage avec moyenne mobile
    x_opposite_ankle = df[side_params['opposite_side_displacement']].rolling(window=5).mean()

    # Trouver les pics (max) et les vallées (min)
    peaks, _ = find_peaks(x_opposite_ankle)  # Trouver les pics
    valleys, _ = find_peaks(-x_opposite_ankle)  # Trouver les vallées en inversant le signal

    displacements = []
    frame_intervals = []  # Pour stocker les intervalles de frame pour chaque déplacement

    # Calculer la différence entre chaque pic et la vallée qui le précède
    for peak in peaks:
        preceding_valleys = valleys[valleys < peak]  # Trouver les vallées avant ce pic
        if len(preceding_valleys) > 0:
            last_valley = preceding_valleys[-1]
            displacement = x_opposite_ankle[peak] - x_opposite_ankle[last_valley]
            # Condition pour n'ajouter le déplacement que s'il est supérieur à 90
            if displacement > 90:
                displacements.append(displacement)
                frame_intervals.append(f"{last_valley}-{peak}")  # Ajout de l'intervalle de frames

    # Retourne les déplacements validés et leurs intervalles de frame
    return pd.DataFrame({'Displacement': displacements, 'Frame_Interval': frame_intervals})



# Fonction pour convertir le DataFrame en format Excel
def save_to_excel(df, file_path):
    # Remplacer les séparateurs de milliers et de décimales
    df = df.map(lambda x: f"{x:,.2f}".replace(",", " ").replace(".", ".") if isinstance(x, (int, float)) else x)

    # Créer un classeur Excel
    wb = Workbook()
    ws = wb.active

    # Écrire les en-têtes
    for c_idx, column in enumerate(df.columns, 1):
        ws.cell(row=1, column=c_idx, value=column)

    # Écrire les données dans le classeur
    for r_idx, row in enumerate(df.values, 2):  # Commence à la ligne 2 pour les données
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Enregistrer le fichier
    wb.save(file_path)









# Charger le fichier CSV d'entrée
input_file = (f'D:\\Documents\\UCL\\Master1\\Techno\\TP1\\{params[side_selector]["Side_name"]}'
              f'\\{params[side_selector]["Side_name"]}.csv')
df = pd.read_csv(input_file)

# Calculer les angles
angles_df = calculate_angles(df, params[side_selector])

output_file = f'D:\\Documents\\UCL\\Master1\\Techno\\TP1\\{params[side_selector]["Side_name"]}\\'

# Tracer et enregistrer les graphes des angles dans des fichiers PNG séparés
if angles_df is not None:
    plot_angle(angles_df['Knee_Angle'], "Knee Angle", f"{output_file}knee_angle.png")
    plot_angle(angles_df['Hip_Angle'], "Hip Angle", f"{output_file}hip_angle.png")
    plot_angle(angles_df['Ankle_Angle'], "Ankle Angle", f"{output_file}ankle_angle.png")
else:
    print("Les angles n'ont pas pu être calculés en raison de colonnes manquantes.")

# Tracer et enregistrer le déplacement de la cheville en x
plot_displacement(df, f"{output_file}ankle_displacement.png", params[side_selector])

# Calculer le déplacement horizontal du pied
displacement_df = calculate_horizontal_displacement(df, params[side_selector])

# Enregistrer les résultats dans des fichiers Excel
save_to_excel(angles_df, f'{output_file}output_angles.xlsx')
save_to_excel(displacement_df, f'{output_file}output_displacement.xlsx')

print("Angles calculés, graphiques sauvegardés et déplacements enregistrés.")
